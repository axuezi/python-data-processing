from openpyxl import load_workbook
from config.public_data import excel_meter_reading_path
from utils.mysql_util import MysqlUtil

workbook = load_workbook(filename=excel_meter_reading_path)
ws = workbook["Sheet1"]

db = MysqlUtil()


def save(no, device_code, meter_type, read_date, time_period, this_value, indication):
    # 保存抄表数据
    sql = "INSERT INTO saas_solution_meter_reading.t_meter_device_time_interval_statistics (id, tenement_code, " \
          "project_code, device_code, meter_type, read_date, last_date, time_period, indication, magnification, " \
          "`last_value`, this_value, version, create_time, create_by, update_time, update_by, logic_del) " \
          "VALUES ('%s', 'ZH_00001', 'ZH_00001_XM_00000001', '%s','%s', '%s', null, '%s', '%s', 1, null, '%s'," \
          " 1, '2021-08-18 10:00:24', '','2021-08-18 10:00:24', '', 0) " \
          % (no, device_code, meter_type, read_date, time_period, indication, this_value)
    db.save(sql)


if __name__ == "__main__":
    a = 1
    for i in ws.iter_rows(min_row=2):
        code = str(i[4].value)
        if code is not None:
            code = "1412341215571628034_" + code
            if i[1].value == '冷量':
                save(a, code, "ColdMeter", str(i[0].value), str(i[2].value), str(i[3].value), str(i[3].value))
            if i[1].value == '电':
                save(a, code, "ElectricityMeter", str(i[0].value), str(i[2].value), str(i[3].value), str(i[3].value))
            if i[1].value == '水表':
                save(a, code, "WaterMeter", str(i[0].value), str(i[2].value), str(i[3].value), str(i[3].value))
            a = a + 1
