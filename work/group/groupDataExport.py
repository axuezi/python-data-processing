from openpyxl import load_workbook
from config.public_data import execl_group_path
from utils.mysql_util import MysqlUtil

workbook = load_workbook(filename=execl_group_path)
ws = workbook["用电月报表D区域"]

db = MysqlUtil()


def selectDeviceByCode(code):
    code = '%' + code + '%'
    sql = "select id from saas_iot_device.t_device where code like '%s'" % code
    return db.selectOne(sql)


def saveGroupDevice(group_id, device_id):
    # 保存分组数据
    sql = "INSERT INTO saas_iot_device.t_group_device (group_id, device_id, tenement_code, project_code, version, " \
          "create_time, create_by, update_time, update_by, logic_del) VALUES (" \
          "'%s', '%s', 'ZH_00001', 'ZH_00001_XM_00000001', 1, '2021-08-20 10:40:02', " \
          "'d5479e28b9094909a1943cfceb9ae270', '2021-08-20 10:40:02', 'd5479e28b9094909a1943cfceb9ae270', 0)" \
          % (group_id, device_id)
    db.save(sql)


def selectGroupList():
    sql = "select id, name from saas_iot_device.t_group where project_code = 'ZH_00001_XM_00000001'"
    return db.selectList(sql)


if __name__ == "__main__":
    for i in ws.iter_rows(min_row=2):
        if i[7].value is not None:
            device_code = str(i[7].value)
            device = selectDeviceByCode(device_code)
            if device is not None:
                deviceId = device['id']
                if deviceId is not None:
                    saveGroupDevice("1427905735858290690", deviceId)
